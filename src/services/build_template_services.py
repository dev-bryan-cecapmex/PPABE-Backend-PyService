import os
from openpyxl import Workbook
from openpyxl.worksheet.datavalidation import DataValidation

class ExcelTemplateService:

    @staticmethod
    def generate_template(catalogos):
        wb = Workbook()

        # Hoja principal
        ws = wb.active
        ws.title = "Beneficiarios"

        # Hoja oculta con catálogos
        ws_cat = wb.create_sheet("Catalogos")

        # Insertar catálogos en hoja oculta
        col = 1
        ranges = {}
        for key, values in catalogos.items():
            ws_cat.cell(1, col, key)
            for i, v in enumerate(values, start=2):
                ws_cat.cell(i, col, v["nombre"])
            # Guardamos rango para validación
            end_row = len(values) + 1
            ranges[key] = f"Catalogos!${chr(64+col)}$2:${chr(64+col)}${end_row}"
            col += 1

        # Definir columnas visibles en hoja Beneficiarios
        headers = []
        for key in catalogos.keys():
            headers.append(key)
            headers.append(f"{key}_ID")

        ws.append(headers)

        # Agregar validaciones
        col_num = 1
        for key in catalogos.keys():
            dv = DataValidation(type="list", formula1=ranges[key], allow_blank=True)
            ws.add_data_validation(dv)
            dv.add(f"{chr(64+col_num)}2:{chr(64+col_num)}1048576")  # toda la columna
            col_num += 2  # saltamos el campo ID

        # Ocultar IDs si la variable está en false
        show_ids = os.getenv("SHOW_IDS", "false").lower() == "true"
        if not show_ids:
            for idx, key in enumerate(catalogos.keys()):
                col_to_hide = (idx * 2) + 2
                ws.column_dimensions[chr(64+col_to_hide)].hidden = True

        # Ocultamos hoja de catálogos
        ws_cat.sheet_state = "hidden"

        return wb
