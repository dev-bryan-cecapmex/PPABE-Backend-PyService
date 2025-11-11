from ..database.connection      import db
from flask                      import jsonify
from config                     import Config

import io 

#Logger
from ..utils.Logger                     import Logger

#Mapeos
from ..utils.Mapeo                      import Mapeo

from ..services.beneficiarios_service   import BeneficiariosService
from ..services.contacto_service        import ContactosService 
from ..services.apoyo_service           import ApoyosService
from ..services.search_service          import SearchService

from datetime                           import datetime

import traceback

import polars as pl 
import uuid

import re
from datetime import datetime

import os
from openpyxl import Workbook
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


class ExcelService:  
    
    @staticmethod
    def process_file(file, id_user, id_dependencia_user):
        try:
            Logger.add_to_log("info", "="*30)
            Logger.add_to_log("info", f"INICIO DE CARGA MASIVA")
            Logger.add_to_log("info", "="*30)
            
            Logger.add_to_log("info", f"Id User: {id_user}")
            Logger.add_to_log("info",f"Dependencia:{id_dependencia_user}")
            
            #data = pl.read_excel(io.BytesIO(file.read()))
            #Logger.add_to_log("info", data)
            #data = data.filter(~pl.all_horizontal(pl.all().is_null()))
            data = pl.read_excel(
                    io.BytesIO(file.read()),
                    schema_overrides = Config.CELLS_DATA_TYPES,
                    infer_schema_length=10000
                )

            data = data.with_columns(
                
                pl.col("Fecha de Nacimiento").is_null().alias("fecha_nac_vacia_original"),
                
                pl.col("Fecha de Nacimiento")
                .str.strptime(pl.Datetime, "%Y-%m-%d %H:%M:%S", strict=False)
                .dt.strftime("%d/%m/%Y")
                .alias("Fecha de Nacimiento")
            )
            
            data = data.with_columns([
                    pl.col("Estado Civil").is_null().alias("estado_civil_vacio_original"),
                    pl.col("Sexo").is_null().alias("sexo_vacio_original"),
                   
                ])
                        
            # También eliminar filas que estén vacías o solo tengan espacios
            data = data.filter(
                pl.any_horizontal(
                    pl.when(pl.col(c).is_not_null() & (pl.col(c).cast(pl.Utf8).str.strip_chars() != ""))
                    .then(True)
                    .otherwise(False)
                    for c in data.columns
                )
            )
            
            rows = data.to_dicts()
            
            Logger.add_to_log("info", "Columnas de los datos")
            Logger.add_to_log("info", data.columns)
            
            Logger.add_to_log("info", f"Total de filas del Excel: {len(rows)}")
            
            Logger.add_to_log("info", f"Inicio de estrucutura de los datos .....")
            
            # Listado de Beneficiarios Nuevos para insertar en BD
            beneficiarios_to_insert = []
            
            # Set de IDs de beneficiarios nuevos 
            beneficiarios_nuevos_ids = set()
            
            # Lista de relacciones completas: fila -> beneficiario -> contacto -> apoyo
            relaciones = []
            
            # Lista de filas con errores de validacion
            rows_errors = []
            
            """
                CACHE LOCAL: Detecta duplicidad DENTRO del Excel
                Estrucutura: {(curp, rfc): id_beneficiario}
            """
            cache_beneficiarios_excel = {}
            
            Logger.add_to_log("info", f"Inicio de estrucutura correctamente")
            
            Logger.add_to_log("info", "Extrayendo grupos de columnas")
            
            # INICIO de agrupamiento
            
            # GRUPO 1: Columna de Beneficiarios
            group_one_df = data.select(Config.GROUP_ONE_KEYS).to_dict()
           
           
            # GRUPO 2: Columnas de Contacto
            group_two_df = data.select(Config.GROUP_TWO_KEYS).to_dict()
           
           
            # GRUPO 3: Columnas de Apoyos
            group_tree_df = data.select(Config.GROUP_TREE_KEYS).to_dict()
            
            
            # FIN de agrupamiento
            
            # MAPEO por Grupos
            # Grupo 1 - Beneficiarios
            sexos_map = SearchService.get_sexo_map()
          
            
            # Grupo 2 - Contactos
            estados_map = SearchService.get_estado_map()
            municipios_map = SearchService.get_municipio_map()
            colonias_map = SearchService.get_colonia_map()
            estados_civiles_map = SearchService.get_estado_civil_map()
        
            # Grupo 3 - Apoyos
            dependencias_map = SearchService.get_dependencias_map()
            programas_map = SearchService.get_programas_map()
            subprograma_map = SearchService.get_subprogramas_map()
            componentes_map = SearchService.get_componentes_map()
            acciones_map = SearchService.get_acciones_map()
            tipos_beneficiarios_map = SearchService.get_tipos_beneficiarios_map()
            
            # Mapa de beneficiarios existentes en BD (para detectar duplicados con BD)
            beneficiario_map = SearchService.get_beneficiarios_map()
            Logger.add_to_log("info", f"  ✓ Beneficiarios existentes en BD: {len(beneficiario_map)} registros")
        
            # Carpeta de Beneficiarios 
            carpetas_beneficiarios_map = SearchService.get_carpeta_beneficiarios_map()
            Logger.add_to_log("info", f"  ✓ Carpetas Beneficiarios: {len(carpetas_beneficiarios_map)} registros")
            Logger.add_to_log("info", "✓ Todos los catálogos cargados exitosamente")
            
            # Diccionario de Estadistica
            stats = {
                'total_filas': len(rows),
                'beneficiarios_nuevos': 0,
                'beneficiarios_existentes_db': 0,
                'duplicados_en_excel': 0,
                'errores_validacion': 0
            }

            for idx, row in enumerate(rows):
                curp = row.get('Curp') or None
                rfc = row.get('RFC') or None

                # Quitar los espacios
                if curp:
                    curp = curp.strip()
                if rfc:
                    rfc = rfc.strip()

                # ==============================
                # Grupo 1 - Beneficiarios
                # ==============================
                sexo = row.get('Sexo')
                id_sexo = sexos_map.get(sexo.upper().rstrip()) if sexo else None

                # ==============================
                # Grupo 2 - Contacto
                # ==============================
                calle = row.get('Calle')
                numero = row.get('Numero')

                estado = row.get('Estado (catálogo)')
                id_estado = estados_map.get(estado.upper().rstrip()) if estado else None

                municipio = row.get('Municipio Dirección (catálogo)')
                id_municipio = municipios_map.get(municipio.upper().rstrip()) if municipio else None

                estado_civil = row.get('Estado Civil')
                id_estado_civil = estados_civiles_map.get(estado_civil.upper().rstrip()) if estado_civil else None

                telefono = row.get('Telefono')
                telefono_2 = row.get('Telefono 2')
                correo = row.get('Correo')
                monto = row.get('Monto')

                colonia = row.get('Colonia')
                colonia = colonia.upper().rstrip() if colonia else None

                # ==============================
                # Grupo 3 - Apoyos
                # ==============================
                dependencia = row.get('Dependencia')
                id_dependencia = dependencias_map.get(dependencia.upper().rstrip()) if dependencia else None

                programa = row.get('Programa')
                id_programa = programas_map.get((programa.upper().rstrip(), id_dependencia)) if programa and id_dependencia else None

                subprograma = row.get('Subprograma')
                id_subprograma = subprograma_map.get((subprograma.upper().rstrip(), id_programa)) if subprograma and id_programa else None

                componente = row.get('Componente')
                id_componente = componentes_map.get((componente.upper().rstrip(), id_subprograma)) if componente and id_subprograma else None

                accion = row.get('Accion')
                id_acciones = acciones_map.get(accion.upper().rstrip()) if accion else None

                tipo_beneficio = row.get('Tipo de Beneficio')
                id_tipo_beneficiario = tipos_beneficiarios_map.get(tipo_beneficio.upper().rstrip()) if tipo_beneficio else None

                
                # Carpeta Beneficiario
                
                fecha_plantilla= row.get('Fecha de Registro')
                row['Fecha de Nacimiento'] = datetime.strptime(
                row['Fecha de Nacimiento'], 
                    '%d/%m/%Y'
                ).strftime('%Y-%m-%d')
                
                fecha_nacimiento = row['Fecha de Nacimiento'] 
            
            
                # Valicacciones
                validacion_errores = {}
                msg_error = ""
                
                fecha = None
                
                if not fecha_plantilla:
                    validacion_errores['Fecha de Registro'] = 'Celda vacía'
                else:
                    if isinstance(fecha_plantilla, str):
                        # Plantilla del sistema
                        fecha_str = fecha_plantilla
                        fecha = datetime.strptime(fecha_str, "%d/%m/%Y" )
                        
                    else:
                        # Plantilla Ruy
                        fecha = fecha_plantilla
                       
                
                
                if fecha:
                    mes = fecha.month
                    anio = fecha.year
                   
                    id_carpeta_beneficiario = carpetas_beneficiarios_map.get((mes, anio, id_dependencia_user))
                    Logger.add_to_log("info", f"Carpeta Beneficiario: {id_carpeta_beneficiario}")
                else:
                    validacion_errores['Fecha de Registro'] = 'Error en formato'
                    
                if not id_carpeta_beneficiario:
                    validacion_errores['Carpeta de Beneficiarios'] = f"No existe carpeta para {mes}/{anio}"
                
                
                if (len(curp or '') > 18 or len(curp or '') < 18 ) and curp != None:
                    validacion_errores['Curp'] = row.get('Curp')
                    msg_error = "Curp inválida. Debe tener 18 caracteres."

                if not fecha_nacimiento :
                    if not row["fecha_nac_vacia_original"]:
                        validacion_errores['Fecha de Nacimiento'] = 'Error en formato'
                
                if not id_sexo :
                    if not row["sexo_vacio_original"]:
                        validacion_errores['Sexo'] = row.get('Sexo')
                        
                if calle == None or calle.strip() == "":
                    validacion_errores['Calle'] = 'Celda vacía'
                
                if numero == None or numero.strip() == "":
                    validacion_errores['Número'] = 'Celda vacía'
                
                if not id_estado_civil:
                    if not row["estado_civil_vacio_original"]:
                        validacion_errores['Estado Civil'] = row.get('Estado Civil')
                        
                if not id_estado:
                    validacion_errores['Estado'] = row.get('Estado (catálogo)')
                    
                if not id_municipio:
                    validacion_errores['Municipio'] = row.get('Municipio Dirección (catálogo)')
                    
                if not colonia:
                    validacion_errores['Colonia'] = 'Celda vacía'
                    
                if not telefono:
                    validacion_errores['Telefono'] = 'Celda vacía'
                
                if not telefono_2:
                    validacion_errores['Telefono 2'] = 'Celda vacía'
                
                if not correo:
                    validacion_errores['Correo'] = 'Celda vacía'
                
                if not monto:
                    validacion_errores['Monto'] = 'Celda vacía'
                
                if not id_tipo_beneficiario:
                    validacion_errores['Tipo de Beneficio'] = row.get('Tipo de Beneficio')
                
                if not id_dependencia:
                    validacion_errores['Dependecia'] = row.get('Dependencia')
               
                if not id_programa:
                    validacion_errores['Programa'] = row.get('Programa')    
                    
                if not id_subprograma:
                     validacion_errores['Subprograma'] = row.get('Subprograma')   

                if not id_componente:
                    validacion_errores['Componente'] = row.get('Componente')
                
                if not id_acciones:
                    validacion_errores['Accion'] = row.get('Accion')
                
                if validacion_errores:
                    stats['errores_validacion'] += 1
                    for validador in validacion_errores:
                    
                        error_detail = {
                            'row_index': idx + 2,
                            'curp': row.get('Curp'),
                            'nombre_completo': f"{row.get('Nombre', '')} {('Apellido paterno', '')} {row.get('Apellido Materno', '')}".strip(),
                            'error': msg_error or 'Error de validación en campos obligatorios',
                            'campos_invalidos': validador,
                            'valor': validacion_errores[validador],
                            'data': row
                        }
                        
                        Logger.add_to_log("info","Errores fatales")
                        rows_errors.append(error_detail)
                    
                   
                    continue
                
                
                    
                


                id_beneficiario = None
                es_nuevo = False
                origen = "" # Para Tracking: 'cache_excel', 'db', 'nuevo'
                
                # Buscar en CHACHE LOCAL del Excel
                if curp or rfc:
                    key_beneficiario = (curp, rfc)
                    
                if key_beneficiario in cache_beneficiarios_excel:
                    id_beneficiario = cache_beneficiarios_excel[key_beneficiario]
                    stats['duplicados_en_excel'] +=1
                    origen = 'cache_excel'
                   
                # Busqueda por solo CURP en cache
                elif curp and not id_beneficiario:
                    for(c,r), id_ben in cache_beneficiarios_excel.items():
                        if c == curp:
                            id_beneficiario = id_ben
                            stats['duplicados_en_excel'] += 1
                            origen = "cache_excel"
                           
                            break
                            
                # Busqueda por solo RFC en cache  
                elif rfc and not id_beneficiario:
                    for(c,r), id_ben in cache_beneficiarios_excel.items():
                        stats['duplicados_en_excel'] += 1
                        origen = "cache_excel"
                        
                        break
                    
                # Busqueda en Base de Datos
                if not id_beneficiario:
                    # Busqueda por CURP y RFC
                    if curp and rfc:
                        id_beneficiario = beneficiario_map.get((curp, rfc))
                        if id_beneficiario:
                            origen = 'db'
                            
                    # Busqueda solo por CURP
                    elif curp and not id_beneficiario:
                        id_beneficiario = next(
                            (id_ben for (c,_), id_ben in beneficiario_map.items() if c == curp),
                            None
                        )
                        if id_beneficiario:
                            origen ='db'
                            
                    # Busqueda solo por RFC
                    elif rfc and not id_beneficiario:
                        id_beneficiario = next(
                            (id_ben for (_,r), id_ben in beneficiario_map.items() if r == rfc),
                            None
                        )
                        if id_beneficiario:
                            origen = 'db'
                    
                    if id_beneficiario and origen == 'db':
                        stats['beneficiarios_existentes_db'] += 1
                       
                # Crea Nuevo beneficiario
                if not id_beneficiario:
                    id_beneficiario = str(uuid.uuid4())
                    es_nuevo = True
                    origen = "nuevo"
                    stats['beneficiarios_nuevos'] += 1
                    beneficiarios_nuevos_ids.add(id_beneficiario)
                    
                    # Objeto con beneficiario con mapeo correcto
                    nuevo_beneficiario = {
                        'id': id_beneficiario,
                        'creador': id_user,
                        'modificador': id_user,
                    }
                    
                    # Mapeo columnas del Excel a columnas de BD
                    for excel_col in Config.GROUP_ONE_KEYS:
                        db_col = Config.COLUMN_MAP_GROUP_ONE.get(excel_col, excel_col)
                        nuevo_beneficiario[db_col] = row.get((excel_col))
                    
                    # Asegurar que tenga el idSexo correcto
                    nuevo_beneficiario['idSexo'] = id_sexo

                    # Agregar a lista de inserción 
                    beneficiarios_to_insert.append(nuevo_beneficiario) 
                    
                    # REGISTRO en CACHE LOCAL
                    if curp or rfc:
                        cache_beneficiarios_excel[(curp, rfc)] = id_beneficiario
                                       
                    
                # Preparacion de contacto y apoyo
                
                # Pre-generar UUIDs temporales
                id_contacto_temp    = str(uuid.uuid4())
                id_apoyo_temp       = str(uuid.uuid4())
                
                # Construccion de objeto de CONTACTO
                contacto_data = {
                    'id': id_contacto_temp,
                    'creador': id_user,
                    'modificador': id_user,
                }
                
                for excel_col in Config.GROUP_TWO_KEYS:
                    if excel_col in Config.COLUMN_MAP_GROUP_TWO:
                        db_col = Config.COLUMN_MAP_GROUP_TWO[excel_col]
                        contacto_data[db_col] = row.get(excel_col)
                
                # Agregar ID's de catálogos
                contacto_data['idEstado']       = id_estado
                contacto_data['idMunicipio']    = str(id_municipio[1]) if id_municipio else None
                contacto_data['colonia']        = colonia if colonia else None
                contacto_data['idEstadoCivil']  = id_estado_civil
                
                
                # Construccion de objeto de APOYO
                apoyo_data = {
                    'id': id_apoyo_temp,
                    'idBeneficiario': id_beneficiario,
                    'idContacto': id_contacto_temp,
                    'creador': id_user,
                    'modificador': id_user,
                }
                
                # Mapear columnas del Excel a columnas de DB
                for excel_col in Config.GROUP_TREE_KEYS:
                    db_col = Config.COLUMN_MAP_GROUP_TREE.get(excel_col, excel_col)
                    apoyo_data[db_col] = row.get(excel_col)
                
                # Agregar IDs de catálogos
                apoyo_data['idDependencia']     = id_dependencia
                apoyo_data['idPrograma']        = id_programa
                apoyo_data['idSubprograma']     = id_subprograma
                apoyo_data['idComponente']      = id_componente
                apoyo_data['idAccion']          = id_acciones
                apoyo_data['idTipoBeneficio']   = id_tipo_beneficiario  
                # Agregar despues idCarpetaBeneficiarios
                apoyo_data['idCarpetaBeneficiarios'] = id_carpeta_beneficiario
                # Registro de relación completa
                relacion = {
                    'row_index': idx + 2,
                    'id_beneficiario': id_beneficiario,
                    'id_contacto': id_contacto_temp,
                    'id_apoyo': id_apoyo_temp,
                    'es_beneficiario_nuevo': es_nuevo,
                    'origen_beneficiario': origen,
                    'contacto_data': contacto_data,
                    'apoyo_data': apoyo_data,
                    # Datos para reporte
                    'curp': curp,
                    'rfc': rfc,
                    'nombre_completo': f"{row.get('Nombre',' ')} {row.get('Apellido paterno','')} {row.get('Apellido Materno','')}".strip()
                }
                
                relaciones.append(relacion)
                 
                
                # Fin de loop
            # Estadistica y reporte de duplicados
            Logger.add_to_log("info", "")
            Logger.add_to_log("info", "=" * 60)
            Logger.add_to_log("info", "📊 ESTADÍSTICAS DE FASE 1:")
            Logger.add_to_log("info", "=" * 60)
            Logger.add_to_log("info", f"  Total de filas procesadas: {stats['total_filas']}")
            Logger.add_to_log("info", f"  ✨ Beneficiarios NUEVOS: {stats['beneficiarios_nuevos']}")
            Logger.add_to_log("info", f"  ✓ Beneficiarios EXISTENTES en BD: {stats['beneficiarios_existentes_db']}")
            Logger.add_to_log("info", f"  ♻️  Duplicados EN EXCEL: {stats['duplicados_en_excel']}")
            Logger.add_to_log("info", f"  ⚠️  Errores de validación: {stats['errores_validacion']}")
            Logger.add_to_log("info", f"  📝 Relaciones válidas creadas: {len(relaciones)}")
            Logger.add_to_log("info", "=" * 60)
            Logger.add_to_log("info", "")
            
            # Reporte detallado de duplicados en Excel
            if stats['duplicados_en_excel'] > 0:
    
                # Contar ocurrencias de cada beneficiario
                beneficiario_ocurrencias = {}
                for rel in relaciones:
                    id_ben = rel['id_beneficiario']
                    if id_ben not in beneficiario_ocurrencias:
                        beneficiario_ocurrencias[id_ben] = {
                            'count': 0,
                            'curp': rel['curp'],
                            'rfc': rel['rfc'],
                            'nombre': rel['nombre_completo'],
                            'filas': []
                        }
                    beneficiario_ocurrencias[id_ben]['count'] += 1
                    beneficiario_ocurrencias[id_ben]['filas'].append(rel['row_index'])
                
                # Filtrar solo los que aparecen más de una vez
                duplicados = {k: v for k, v in beneficiario_ocurrencias.items() if v['count'] > 1}
                """ 
                for id_ben, info in duplicados.items():
                    Logger.add_to_log("warn", f"  • {info['nombre']}")
                    Logger.add_to_log("warn", f"    CURP: {info['curp']}, RFC: {info['rfc']}")
                    Logger.add_to_log("warn", f"    Aparece {info['count']} veces en filas: {info['filas']}")
                    Logger.add_to_log("warn", "")
                """
            
            # Reporte de errores de validación 
            if rows_errors:
                Logger.add_to_log("error", "REPORTE DE ERRORES DE VALIDACIÓN")
                Logger.add_to_log("error","-" * 60)
                
                for error in rows_errors[:10]: # Muesta solo primero 10
                    Logger.add_to_log("error", f"  Fila {error['row_index']}: {error.get('nombre_completo', 'N/A')}")
                    Logger.add_to_log("error", f"    CURP: {error.get('curp', 'N/A')}")
                    Logger.add_to_log("error", f"    Error: {error['error']}")
                    Logger.add_to_log("error", f"    Campos inválidos: {error['campos_invalidos']}")
                    Logger.add_to_log("error", "")
                
                if len(rows_errors) > 10:
                    Logger.add_to_log("error", f"  ... y {len(rows_errors) - 10} errores más")
                
                return jsonify({
                    'success':False,
                    'message':'No se encontraron registros validos para procesar',
                     'data':{
                        'total_filas':stats['total_filas'],
                        'errores': stats['errores_validacion'],
                        'errores_detalle': rows_errors
                    },
                    'error':'Sin datos válidos'
                }),400


            
            Logger.add_to_log("info","INICIANDO INSERCIÓN EN BASE DE DATOS ...")
            
            if not relaciones:
                Logger.add_to_log("warn", "No hay datos validos para insertar")
                return jsonify({
                    'success':False,
                    'message':'No se encontraron registros validos para procesar',
                     'data':{
                        'total_filas':stats['total_filas'],
                        'errores': stats['errores_validacion'],
                        'errores_detalle': rows_errors
                    },
                    'error':'Sin datos válidos'
                }),400
                
            # Insercion de beneficiarios nuevos
            if beneficiarios_to_insert:
                try:
                    Logger.add_to_log('info', f"💾 🗄️ Insertando {len(beneficiarios_to_insert)} beneficiarios nuevos ...")
                    Logger.add_to_log('debug', f"Primeros 3 beneficiarios: { beneficiarios_to_insert[:3]}")
                    # Llamada de al servicio de insercion
                    BeneficiariosService.bulk_insert(beneficiarios_to_insert)
                    Logger.add_to_log('info', f"✅ 💾 {len(beneficiarios_to_insert)} beneficiarios insertados exitosamente")
                except Exception as e:  
                    Logger.add_to_log('error', "❌ 💾 ERROR AL INSERTAR BENEFICIARIOS")
                    Logger.add_to_log('error', f"Detalles: {str(e)}")      
                    Logger.add_to_log("error", traceback.format_exc())
                    
                    return jsonify({
                        'success':False,
                        'message':'Error al insertar beneficiarios',
                        'data':{
                            'fase_fallida':'Insercion de Beneficiarios',
                            'beneficiarios_intentados': len(beneficiarios_to_insert)
                        },
                        'error':str(e)
                    }), 500       
            else:
                Logger.add_to_log("info", "✅ 💾 No hay beneficiarios nuevos para insertar")
                     
            # Preparacion de Contactos y Apoyos
            Logger.add_to_log('info', "Preparando lista de contactos y apoyos ...")
            
            contactos_to_insert = []
            apoyos_to_insert    = []
            
            for relacion in relaciones:
                # Extraer datos ya mapedas
                contactos_to_insert.append(relacion['contacto_data'])
                apoyos_to_insert.append(relacion['apoyo_data'])
            
            Logger.add_to_log("info", f"{len(contactos_to_insert)} contactos preparados")
            Logger.add_to_log("info", f"{len(apoyos_to_insert)} apoyos preparados")
            
            # INSERCIÓN DE CONTACTOS
            if contactos_to_insert:
                try:
                    Logger.add_to_log("info", f"💾 🗄️ Insertando {len(contactos_to_insert)} contactos nuevos ...")
                    Logger.add_to_log('debug', f"Primeros 3 contactos: {contactos_to_insert[:3]}")
                    # Llamada de al servicio de insercion
                    ContactosService.bulk_insert(contactos_to_insert)
                    Logger.add_to_log('info', f"✅ 💾 {len(contactos_to_insert)} contactos insertados exitosamente")

                except Exception as e:
                    Logger.add_to_log("error", "❌ 💾 ERROR AL INSERTAR CONTACTOS")
                    Logger.add_to_log("error", f"Detalles: {str(e)}")
                    Logger.add_to_log("error", traceback.format_exc())
                    
                    return jsonify({
                        'success':False,
                        'message':'Error al insertar contactos',
                        'data':{
                            'fase_fallida':'Insercion de contactos',
                            'beneficiarios_insertados': len(beneficiarios_to_insert),
                            'contactos_intentados': len(contactos_to_insert),
                            'warning':'Los beneficiarios quedaron en BD sin contactos asociados'
                        },
                        'error':str(e)
                    }), 500   
            else:
                Logger.add_to_log("warn", "✅ 💾 No hay contactos para insertar")       
            
            # INSERCIÓN DE APOYOS
            if apoyos_to_insert:
                try:
                    Logger.add_to_log("info", f"💾 🗄️ Insertando {len(apoyos_to_insert)} contactos nuevos ...")
                    Logger.add_to_log('debug', f"Primeros 3 apoyos: {apoyos_to_insert[:3]}")
                    # Llamada de al servicio de insercion
                    ApoyosService.bulk_insert(apoyos_to_insert)
                    Logger.add_to_log('info', f"✅ 💾 {len(apoyos_to_insert)} apoyos insertados exitosamente")

                except Exception as e:
                    Logger.add_to_log("error", "❌ 💾 ERROR AL INSERTAR APOYOS")
                    Logger.add_to_log("error", f"Detalles: {str(e)}")
                    Logger.add_to_log("error", traceback.format_exc())
                    
                    return jsonify({
                        'success':False,
                        'message':'Error al insertar apoyos',
                        'data':{
                            'fase_fallida':'Insercion de apoyos',
                            'beneficiarios_insertados': len(beneficiarios_to_insert),
                            'contactos_intentados': len(contactos_to_insert),
                            'apoyos_intentados': len(apoyos_to_insert),
                            'warning':'Los beneficiarios quedaron en BD sin contactos asociados'
                        },
                        'error':str(e)
                    }), 500   
            else:
                Logger.add_to_log("warn", "✅ 💾 No hay contactos para insertar")       
            
            
            
            
        except Exception as ex:
            
            return jsonify({
            'success': False,
            'message': 'Error crítico en el proceso de carga masiva',
            'data': None,
            'error': {
                'type': type(ex).__name__,
                'message': str(ex),
                'traceback': traceback.format_exc()
            }
        }), 500 
       
    @staticmethod
    def generate_template(catalogos):
        wb = Workbook()
        ws = wb.active
        ws.title = "Beneficiarios"

        # ---------- Hoja oculta con catálogos ----------
        ws_cat = wb.create_sheet("Catalogos")
        col = 1

        # Rango para DataValidation (con "=") y rangos crudos para fórmulas
        dv_ranges = {}       # p.ej. "=Catalogos!$A$2:$A$100"
        lookup_ranges = {}   # p.ej. ("Catalogos!$A$2:$A$100","Catalogos!$B$2:$B$100","Catalogos!$A$2:$B$100")

        for key, values in catalogos.items():
            if not values:
                continue

            # Columna de nombres
            name_col_letter = get_column_letter(col)
            # Columna de IDs (adyacente)
            id_col_letter = get_column_letter(col + 1)

            ws_cat.cell(1, col, key)
            ws_cat.cell(1, col + 1, f"{key}_ID")

            for i, v in enumerate(values, start=2):
                ws_cat.cell(i, col, v["nombre"])
                ws_cat.cell(i, col + 1, v["id"])

            end_row = len(values) + 1

            # Para DV (debe llevar "=")
            dv_ranges[key] = f"=Catalogos!${name_col_letter}$2:${name_col_letter}${end_row}"

            # Para fórmulas (sin "=")
            names_raw = f"Catalogos!${name_col_letter}$2:${name_col_letter}${end_row}"
            ids_raw   = f"Catalogos!${id_col_letter}$2:${id_col_letter}${end_row}"
            table_raw = f"Catalogos!${name_col_letter}$2:${id_col_letter}${end_row}"
            lookup_ranges[key] = (names_raw, ids_raw, table_raw)

            col += 2

        # ---------- Encabezados visibles (una sola fila) ----------
        headers = [
            "Curp", "Nombre", "Apellido paterno", "Apellido Materno",
            "Fecha de Nacimiento", "Estado (catálogo)", "Estado Civil", "Sexo",
            "Calle", "Numero", "Colonia", "Municipio Dirección (catálogo)",
            "Telefono", "Telefono 2", "Correo", "Programa", "Componente",
            "Accion", "Fecha de Registro", "Monto", "Tipo de Beneficio",
            "RFC", "Regimen Capital", "Actividad", "Nombre Comercial",
            "Razón Social", "Localidad", "Dependencia", "Subprograma"
        ]
        ws.append(headers)

        # Formato de cabecera
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        border_style = Side(style="thin", color="000000")
        for c in range(1, len(headers) + 1):
            cell = ws.cell(1, c)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = Border(top=border_style, bottom=border_style, left=border_style, right=border_style)
            ws.column_dimensions[get_column_letter(c)].width = 18
        ws.freeze_panes = "A2"

        # Campos que llevan listas (catálogos)
        catalog_fields = {
            "Estado (catálogo)": "Estado",
            "Municipio Dirección (catálogo)": "Municipio",
            "Sexo": "Sexo",
            "Estado Civil": "EstadoCivil",
            "Programa": "Programa",
            "Componente": "Componente",
            "Accion": "Accion",
            "Tipo de Beneficio": "TipoBeneficio",
            "Dependencia": "Dependencia",
            "Colonia": "Colonia"
        }

        # ---------- Data Validation (listas) ----------
        MAX_ROWS = 10000
        for idx, h in enumerate(headers, start=1):
            if h in catalog_fields:
                key = catalog_fields[h]
                if key in dv_ranges:
                    col_letter = get_column_letter(idx)
                    dv = DataValidation(type="list", formula1=dv_ranges[key], allow_blank=True)
                    ws.add_data_validation(dv)
                    dv.add(f"{col_letter}2:{col_letter}{MAX_ROWS + 1}")

        # ---------- Agregar columnas ID (al final) ----------
        id_headers = [f"{catalog_fields[h]}_ID" for h in headers if h in catalog_fields]
        start_id_col = len(headers) + 1
        for i, idh in enumerate(id_headers):
            ws.cell(1, start_id_col + i, idh)
            cell = ws.cell(1, start_id_col + i)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = Border(top=border_style, bottom=border_style, left=border_style, right=border_style)
            ws.column_dimensions[get_column_letter(start_id_col + i)].width = 18

        # ---------- Fórmulas automáticas para _ID (XLOOKUP + fallback VLOOKUP) ----------
        # Para cada campo con catálogo, ubicamos columna de nombre y columna de ID
        for h in headers:
            if h in catalog_fields:
                cat_key = catalog_fields[h]
                if cat_key in lookup_ranges:
                    names_raw, ids_raw, table_raw = lookup_ranges[cat_key]
                    # Columna del valor seleccionado (catálogo)
                    name_col_idx = headers.index(h) + 1
                    name_col_letter = get_column_letter(name_col_idx)
                    # Columna del ID destino
                    id_header = f"{cat_key}_ID"
                    id_col_idx = start_id_col + id_headers.index(id_header)
                    id_col_letter = get_column_letter(id_col_idx)

                    # Fórmula: intenta XLOOKUP; si no existe, usa VLOOKUP con tabla de 2 columnas contiguas
                    base_formula = (
                        f'IFERROR('
                        f'XLOOKUP({name_col_letter}{{ROW}}, {names_raw}, {ids_raw}, ""),'
                        f'IFERROR(VLOOKUP({name_col_letter}{{ROW}}, {table_raw}, 2, FALSE), "")'
                        f')'
                    )

                    for r in range(2, MAX_ROWS + 2):
                        ws[f"{id_col_letter}{r}"] = f"={base_formula.replace('{ROW}', str(r))}"

        # ---------- Ocultar hoja catálogos ----------
        ws_cat.sheet_state = "hidden"

        # Mostrar/ocultar columnas ID según variable
        show_ids = os.getenv("SHOW_IDS", "false").lower() == "true"
        if not show_ids:
            for i in range(len(id_headers)):
                ws.column_dimensions[get_column_letter(start_id_col + i)].hidden = True

        return wb